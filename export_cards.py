import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from database import get_session_by_id, get_wines_by_session
from database_cards import get_cards_by_session, get_session_cards_summary
from config import EXPORT_DIR


thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _header_style(ws, cols):
    fill = PatternFill(start_color='4F4F4F', end_color='4F4F4F', fill_type='solid')
    font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


def _cell_style(ws, rows, cols):
    for r in range(2, rows + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)


def _autowidth(ws, cols, min_w=14):
    for c in range(1, cols + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = max(max_len + 4, min_w)


SCORE_MAP = {
    1: 'слабо', 2: 'слабо', 3: 'слабо',
    4: 'нормально', 5: 'нормально', 6: 'нормально',
    7: 'хорошо', 8: 'хорошо',
    9: 'очень хорошо', 10: 'очень хорошо',
}


def _score_text(score) -> str:
    if score is None:
        return '-'
    label = SCORE_MAP.get(score, '')
    return f'{score}/10 ({label})'


async def export_cards_excel(session_id: int) -> str:
    session = await get_session_by_id(session_id)
    if not session:
        raise ValueError(f'Session {session_id} not found')

    cards = await get_cards_by_session(session_id)
    summary = await get_session_cards_summary(session_id)

    EXPORT_DIR.mkdir(exist_ok=True)
    safe_title = ''.join(c if c.isalnum() or c in ' _-' else '_' for c in session['title'])
    filename = f"cards_{safe_title}_{session['tasting_date']}.xlsx"
    filepath = EXPORT_DIR / filename

    wb = openpyxl.Workbook()

    # ---- Лист 1: Сводка ----
    ws1 = wb.active
    ws1.title = 'Сводка'
    headers1 = ['Вино', 'Кол-во карточек', 'Средний балл', 'Уровень']
    for i, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=i, value=h)
    _header_style(ws1, len(headers1))

    for i, s in enumerate(summary, 2):
        avg = float(s['avg_score']) if s['avg_score'] else 0
        ws1.cell(row=i, column=1, value=s['wine_name'])
        ws1.cell(row=i, column=2, value=s['card_count'])
        ws1.cell(row=i, column=3, value=round(avg, 2))
        ws1.cell(row=i, column=4, value=SCORE_MAP.get(round(avg), '-'))
    _cell_style(ws1, len(summary) + 1, len(headers1))
    _autowidth(ws1, len(headers1))

    # ---- Лист 2: Все карточки ----
    ws2 = wb.create_sheet('Все карточки')
    headers2 = [
        'Участник', 'Телефон',
        'Вино', 'Позиция',
        'Цвет', 'Аромат', 'Вкус', 'Послевкусие',
        'Дефекты', 'Впечатление', 'Комментарий', 'Оценка',
    ]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=i, value=h)
    _header_style(ws2, len(headers2))

    for i, c in enumerate(cards, 2):
        ws2.cell(row=i, column=1, value=c['participant_name'])
        ws2.cell(row=i, column=2, value=c['participant_phone'])
        ws2.cell(row=i, column=3, value=c['wine_name'])
        ws2.cell(row=i, column=4, value=c['wine_position'])
        ws2.cell(row=i, column=5, value=c['color'] or '')
        ws2.cell(row=i, column=6, value=c['aroma'] or '')
        ws2.cell(row=i, column=7, value=c['taste'] or '')
        ws2.cell(row=i, column=8, value=c['aftertaste'] or '')
        ws2.cell(row=i, column=9, value=c['defects'] or '')
        ws2.cell(row=i, column=10, value=c['impression'] or '')
        ws2.cell(row=i, column=11, value=c['comment'] or '')
        ws2.cell(row=i, column=12, value=_score_text(c['score']))
    _cell_style(ws2, len(cards) + 1, len(headers2))
    _autowidth(ws2, len(headers2))

    # ---- Лист 3: По участникам ----
    ws3 = wb.create_sheet('По участникам')
    row_idx = 1
    if cards:
        seen = set()
        for c in cards:
            pid = c['participant_name']
            if pid in seen:
                continue
            seen.add(pid)

            ws3.cell(row=row_idx, column=1, value=f"{c['participant_name']} ({c['participant_phone']})")
            ws3.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
            row_idx += 1

            sub_headers = ['Вино', 'Цвет', 'Аромат', 'Вкус', 'Послевкусие',
                           'Дефекты', 'Впечатление', 'Комментарий', 'Оценка']
            for j, h in enumerate(sub_headers, 1):
                ws3.cell(row=row_idx, column=j, value=h)
            _header_style(ws3, len(sub_headers))
            row_idx += 1

            for c2 in cards:
                if c2['participant_name'] != pid:
                    continue
                ws3.cell(row=row_idx, column=1, value=c2['wine_name'])
                ws3.cell(row=row_idx, column=2, value=c2['color'] or '')
                ws3.cell(row=row_idx, column=3, value=c2['aroma'] or '')
                ws3.cell(row=row_idx, column=4, value=c2['taste'] or '')
                ws3.cell(row=row_idx, column=5, value=c2['aftertaste'] or '')
                ws3.cell(row=row_idx, column=6, value=c2['defects'] or '')
                ws3.cell(row=row_idx, column=7, value=c2['impression'] or '')
                ws3.cell(row=row_idx, column=8, value=c2['comment'] or '')
                ws3.cell(row=row_idx, column=9, value=_score_text(c2['score']))
                row_idx += 1
            row_idx += 1

    _autowidth(ws3, 9)

    wb.save(str(filepath))
    return str(filepath)